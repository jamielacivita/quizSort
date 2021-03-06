from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk 
from tkinter import filedialog as fd  
import os

DEBUG = False
if __name__ == "__main__":
	DEBUG = False

#load roster data
def load_roster_data():
	return "load roster data"


#load quiz data
def load_quiz_data():
	#import to read xlsx in python.
	#get the 
	#Open PQ.xlsx
	#Switch to Sheet 'Player Level'
	#Read in cell A5
	return "load quiz data"


#generate data student / score / class
def load_data_student_score_class():
	return "load data student score class"
	#if student does not have a score put zero.


#generate csv file(s)
def generate_csv_files():
	return "generate csv files"


#TODO: this is a test funciton - clean it up
def readCellA5():
	#import to read xlsx in python.
	wb = load_workbook(filename = r"c:\source\python\quizSort\Data\PQ.xlsx")
	if DEBUG : print("WB: %s" % wb)
	if DEBUG : print("Sheet Names: %s" % wb.sheetnames)
	#OQ: how to switch to Sheet 'Player Level'
	#ws = wb.get_sheet_by_name("Player Level")
	ws = wb["Player Level"]
	if DEBUG : print("Worksheet : %s " % ws)

	#OQ: How to tead in cell A5

	result = ws['A5'].value
	if DEBUG : print("Result: %s " % result)
	return result


def readCellB2():
	#import to read xlsx in python.
	wb = load_workbook(filename = r"c:\source\python\quizSort\Data\PQ.xlsx")
	if DEBUG : print("In read cell B2 for sheet 1st")
	if DEBUG : print("WB: %s" % wb)
	if DEBUG : print("Sheet Names: %s" % wb.sheetnames)
	#OQ: how to switch to Sheet 'Player Level'
	#ws = wb.get_sheet_by_name("1st")
	ws = wb["1st"]
	if DEBUG : print("Worksheet : %s " % ws)

	#OQ: How to tead in cell A5

	result = ws['B2'].value
	result = result.upper()
	if DEBUG : print("Result: %s " % result)
	return result


def getFirstClassStudents_lst():
	wb = load_workbook(filename = r"c:\source\python\quizSort\Data\PQ.xlsx")
	ws = wb["1st"]
	#Students in the first period class go from B2 to B25
	students_tpl = ws["B2":"B25"]
	students_lst = list(students_tpl)
	students_lst[:] = [s[0].value.upper() for s in students_lst]
	if DEBUG : print("students_lst: %s " % students_lst)
	return students_lst


def writeTable():
	fileName = r"c:\source\python\quizSort\Data\class07.txt"
	print("writing table file.")
	students_lst = getFirstClassStudents_lst()
	with open(fileName, "w") as f:
		for s in students_lst:
			f.write(s + "\n")

def parsePlayers(raw_player):
	if DEBUG : print("In parsePlayers")
	player = raw_player.split("(")[1]
	if DEBUG : print("player:  %s" % player)
	player = player.split(")")[0]
	player = player.strip()
	if DEBUG : print("player: %s" % player)
	return player.upper()

	
def parsePlayerLevelR5():
	#import to read xlsx in python.
	wb = load_workbook(filename = r"c:\source\python\quizSort\Data\PQ.xlsx")
	if DEBUG : print("WB: %s" % wb)
	if DEBUG : print("Sheet Names: %s" % wb.sheetnames)
	ws = wb["Player Level"]
	if DEBUG : print("Worksheet : %s " % ws)
	raw_players = ws['A5'].value
	player = parsePlayers(raw_players)
	raw_score = ws['B5'].value
	score = str(raw_score)
	raw_accuracy = ws['C5'].value
	accuracy = str(int(float(raw_accuracy)*100))
	accuracy = accuracy + "%"
	if DEBUG : print("player: %s " % player)
	if DEBUG : print("score: %s " % score)
	if DEBUG : print("accuracy: %s " % accuracy)
	return (player, score, accuracy)

	

def getParsePlayerLevelRow(workbook, row_number):
	#import to read xlsx in python.
	#wb = load_workbook(filename = r"c:\source\python\quizSort\Data\PQ.xlsx")
	wb = workbook
	if DEBUG : print("WB: %s" % wb)
	if DEBUG : print("Sheet Names: %s" % wb.sheetnames)
	ws = wb["Player Level"]
	if DEBUG : print("Worksheet : %s " % ws)
	raw_players = ws.cell(row = row_number, column=1).value
	player = parsePlayers(raw_players)
	raw_score = ws.cell(row = row_number, column=2).value
	score = str(raw_score)
	raw_accuracy = ws.cell(row = row_number, column = 3).value
	accuracy = str(int(float(raw_accuracy)*100))
	accuracy = accuracy + "%"
	if DEBUG : print("player: %s " % player)
	if DEBUG : print("score: %s " % score)
	if DEBUG : print("accuracy: %s " % accuracy)
	return (player, score, accuracy)


def parsePlayer(wb):
	START_ROW = 5
	END_ROW = 90
	player_level_dict = {}
	for row in range(START_ROW, END_ROW+1):
		key = getParsePlayerLevelRow(wb, row)[0]
		value = getParsePlayerLevelRow(wb, row)
		#check to see if player already has a value in the dictonary.
		if key in player_level_dict:
			old_value = player_level_dict[key]
			if old_value[1] > value[1]:
				continue #because we don't want to add in lower values.
			else:
				player_level_dict[key]=value
		else:	#key is not in dictionary - add it automatically.
			player_level_dict[key]=value
	return player_level_dict


def parseClass(wb, period):

	dir_path = os.path.dirname(os.path.realpath(__file__))
	print(dir_path)
	roster_files_location = dir_path.split("App")[0]+"\\Data\\"

	fileName_dict = {}
	#fileName_dict["01"] = r"c:\source\python\quizSort\Data\class01.txt"
	#fileName_dict["02"] = r"c:\source\python\quizSort\Data\class02.txt"
	#fileName_dict["04"] = r"c:\source\python\quizSort\Data\class04.txt"
	#fileName_dict["07"] = r"c:\source\python\quizSort\Data\class07.txt"
	fileName_dict["01"] = roster_files_location + "class01.txt"
	fileName_dict["02"] = roster_files_location + "class02.txt"
	fileName_dict["04"] = roster_files_location + "class04.txt"
	fileName_dict["07"] = roster_files_location + "class07.txt"

	FILENAME = fileName_dict[period]

	class_scores_out_lst = []
	with open(FILENAME) as f:
		students = f.readlines()
		students[:] = [s.strip() for s in students]
	
	scores_dictionary = parsePlayer(wb)
	for s in students:
		if s in scores_dictionary:
			student = s
			score = scores_dictionary[s][1]
			accuracy = scores_dictionary[s][2]
			class_scores_out_lst.append([student, score, accuracy])
		else:
			student = s
			score = "No Score"
			accuracy = "N/A"
			class_scores_out_lst.append([student, score, accuracy])

	return class_scores_out_lst	


def prettyPrintParseClass(class_scores_out_lst):
	print("{:30}{:20}{:10}".format("STUDENT", "SCORE", "ACCURACY"))
	print("{:30}{:20}{:10}".format("-------", "-----", "--------"))
	for s in class_scores_out_lst:
		student, score, accuracy = s
		print("{:30}{:20}{:10}".format(student, score, accuracy))


def fileDialog():
	print("Path to input file.")
	input_file = input("> ")
	print("Path to output file.")
	output_file = input("> ")

	try:
		with open(input_file) as f:
			f.close()
			return input_file
	except:
		print("Input file does not exist.")
		return None


def startUI():     
	root = tk.Tk()
	input_filename = fd.askopenfilename()     
	output_filename = fd.asksaveasfilename()

	writeOutputNotebook(input_filename, output_filename)
	root.quit()
	root.destroy()

	
	#errmsg = 'Error!' 
	#tk.Button(text='Click to Open File', command=callback).pack(fill=tk.X) 
	root.mainloop()


def writeOutputNotebook(input_filename, output_filename=r"c:\source\python\quizSort\Data\out.xlsx"):
	wb = load_workbook(input_filename)
	
	book = Workbook()
	sheet = book.active

	#create class result sheets in workbook.
	book.create_sheet("Period 01")
	book.create_sheet("Period 02")
	book.create_sheet("Period 04")
	book.create_sheet("Period 07")

	#Do Class Peroid 01
	class01Data = parseClass(wb, "01")
	sheet = book["Period 01"]
	print(book.active)

	for row in range(1,len(class01Data)+1):
		#insert name
		sheet.cell(row = row, column = 1).value = class01Data[row-1][0]
		#insert score
		sheet.cell(row = row, column = 2).value = class01Data[row-1][1]
		#insert score
		sheet.cell(row = row, column = 3).value = class01Data[row-1][2]


	#Do Class Peroid 02
	class02Data = parseClass(wb, "02")
	sheet = book["Period 02"]
	wb.active = 2

	#print(class02Data)

	for row in range(1,len(class02Data)+1):
		#insert name
		sheet.cell(row = row, column = 1).value = class02Data[row-1][0]
		#insert score
		sheet.cell(row = row, column = 2).value = class02Data[row-1][1]
		#insert score
		sheet.cell(row = row, column = 3).value = class02Data[row-1][2]


	#Do Class Peroid 04
	class04Data = parseClass(wb, "04")
	sheet = book["Period 04"]
	wb.active = 3

	#print(class02Data)

	for row in range(1,len(class04Data)+1):
		#insert name
		sheet.cell(row = row, column = 1).value = class04Data[row-1][0]
		#insert score
		sheet.cell(row = row, column = 2).value = class04Data[row-1][1]
		#insert score
		sheet.cell(row = row, column = 3).value = class04Data[row-1][2]


	#Do Class Peroid 07
	class07Data = parseClass(wb, "07")
	sheet = book["Period 07"]
	wb.active = 4

	#print(class02Data)

	for row in range(1,len(class07Data)+1):
		#insert name
		sheet.cell(row = row, column = 1).value = class02Data[row-1][0]
		#insert score
		sheet.cell(row = row, column = 2).value = class02Data[row-1][1]
		#insert score
		sheet.cell(row = row, column = 3).value = class02Data[row-1][2]


	book.save(output_filename)
	return None

def rf():
	#print(os.getcwd())
	dir_path = os.path.dirname(os.path.realpath(__file__))
	print(dir_path)
	roster_files_location = dir_path.split("App")[0]
	roster_files_location = roster_files_location+"Data"
	print(roster_files_location)

def main():
	#print("JWTO")
	#readCellA5()
	#getFirstClassStudents_lst()
	#writeTable()
	#parsePlayerLevelR5()
	#parsePlayerLevel(5)
	
	#print(parsePlayer())
	#prettyPrintParseClass(parseClass("01"))
	#prettyPrintParseClass(parseClass("02"))
	#prettyPrintParseClass(parseClass("04"))
	#prettyPrintParseClass(parseClass("07"))
	#fileDialog()
	startUI()

if __name__ == "__main__":
	main()

